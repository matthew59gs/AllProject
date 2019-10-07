#!/usr/bin/env python
# -*- coding:utf-8 -*-
# public库
# 额外安装包：
# pymysql, openpyxl, dateutil, xlrd, psutil

import os
import sys
import re
import datetime
import logging
import pymysql
import openpyxl
import xlrd
import smtplib
import zipfile
import poplib
import shutil
import dateutil.parser
from email import encoders
from email.header import Header
from email.parser import Parser
from email.header import decode_header
from email.utils import parseaddr
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import pandas as pd
import uuid
from enum import Enum
import psutil
import _ssl

########################################################################################################################
## Enum
########################################################################################################################


class Exchange(Enum):
    SH_A = 1
    SZ_A = 2
    CFFEX = 3
    SHFE = 4
    DCE = 5
    CZCE = 6
    INE = 7


########################################################################################################################
## Other
########################################################################################################################

global logger
logger = None
global MACRO_DATE
MACRO_DATE = "#YYYYMMDD#"
SPLIT_CHAR = "|"

# pd打印显示所有列
pd.set_option('display.max_columns', None)
# pd打印显示所有行
pd.set_option('display.max_rows', None)
# pd打印设置value的显示长度为100，默认为50
pd.set_option('max_colwidth',100)


def SetLog(name):
    path = os.path.join(os.getcwd(), 'logs')
    if not os.path.exists(path):
        os.mkdir(path)
    filename = path + os.sep + name + "_{0}_{1}.log"
    LOG_FORMAT = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    global logger
    logger = logging.getLogger(__name__)
    logger.setLevel(level=logging.DEBUG)
    debug_log = logging.FileHandler(filename.format(
        datetime.datetime.now().strftime("%Y%m%d-%H%M%S"), "DEBUG"))
    debug_log.setLevel(logging.DEBUG)
    debug_log.setFormatter(LOG_FORMAT)
    logger.addHandler(debug_log)
    infolog = logging.FileHandler(filename.format(
        datetime.datetime.now().strftime("%Y%m%d-%H%M%S"), "INFO"))
    infolog.setLevel(logging.INFO)
    infolog.setFormatter(LOG_FORMAT)
    logger.addHandler(infolog)
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)
    console.setFormatter(LOG_FORMAT)
    logger.addHandler(console)
    return logger


def parse_date(_str, normal_diff = 0):
    # 从字符串解析日期
    match = re.findall(r"(\d{8})", _str)
    if match:
        for numbers in match:
            try:
                #解析的日期应该不会差太远，如果超过太远，则
                diff = datetime.datetime.now() - dateutil.parser.parse(numbers)
                if (normal_diff != 0) and ((diff.days > normal_diff) or (diff.days < -1 * normal_diff)):
                    if logger:
                        logger.debug("parse a date[{0}] but over diff[{1}]".format(numbers, normal_diff))
                    continue
                return numbers
            except Exception as e:
                if logger:
                    logger.debug("parse date[{0}] from [{1}] error {2}".format(numbers, _str, e))
    return ""


def Savefile2localpath(data, filename):
    global logger
    if logger:
        logger.info("Save file[{0}]".format(filename))
    try:
        wfile = open(filename, "wb")
        wfile.write(data)
        wfile.close
    except Exception as e:
        if logger:
            logger.error('Save file[{0}] error: {1}'.format(filename, e))


def SingleStringMatch(matchtype, temple, target):
    if ((matchtype == '1' or matchtype == 1) and (target.find(temple) == 0)) or \
        ((matchtype == '2' or matchtype == 2) and (target.find(temple, 1) >= 0)) or \
        ((matchtype == '3' or matchtype == 3) and (target.rfind(temple) == 0)):
        return True
    else:
        return False


def StringMatch(matchtype, temples, target):
    if temples.find(SPLIT_CHAR) >= 0:
        temple_list = temples.split(SPLIT_CHAR)
        for temple in temple_list:
            if SingleStringMatch(matchtype, temple, target):
                return True
    else:
        return SingleStringMatch(matchtype, temples, target)
    return False


def CopyFileFromList(dirlist, filelist):
    if len(dirlist) == 0 or len(filelist) == 0:
        return

    daytime = time.strftime('%Y%m%d')
    for srcdir, destdir in dirlist:
        for filename in filelist:
            a_file = filename.replace(MACRO_DATE, daytime)
            if os.path.isfile(os.path.join(srcdir, a_file)):
                shutil.copy(os.path.join(srcdir, a_file), destdir)


def safe_move_file(src_dir="", src_filename="", target_dir="", target_filename=""):
    return safe_copy_move_file(src_dir=src_dir, src_filename=src_filename,
                               target_dir=target_dir, target_filename=target_filename, mode='move')


def safe_copy_file(src_dir="", src_filename="", target_dir="", target_filename=""):
    return safe_copy_move_file(src_dir=src_dir, src_filename=src_filename,
                               target_dir=target_dir, target_filename=target_filename, mode='copy')


def safe_copy_move_file(src_dir="", src_filename="", target_dir="", target_filename="", i=2, mode=""):
    if os.path.exists(os.path.join(target_dir, target_filename)):
        [pre, suffix] = target_filename.split('.')
        new_filename = pre + "-" + str(i) + '.' + suffix
        if os.path.exists(os.path.join(target_dir, new_filename)):
            safe_copy_move_file(src_dir, src_filename, target_dir, target_filename, i + 1, mode=mode)
        else:
            if mode == 'copy':
                shutil.copy(os.path.join(src_dir, src_filename), os.path.join(target_dir, new_filename))
            elif mode == 'move':
                shutil.move(os.path.join(src_dir, src_filename), os.path.join(target_dir, new_filename))
    else:
        if mode == 'copy':
            shutil.copy(os.path.join(src_dir, src_filename), os.path.join(target_dir, target_filename))
        elif mode == 'move':
            shutil.move(os.path.join(src_dir, src_filename), os.path.join(target_dir, target_filename))
    return True


# 获取交易距离date隔了daydiff天的交易日
# date:		格式YYYYMMDD
# daydiff:	负数，表示向前N天，正数，向后N天，如果是0，且当天是交易日，则返回当天，如果当天不是交易日，与值为1的效果相同
# 返回交易日，如果发生异常，返回0
def GetTradedate(date = datetime.datetime.now().strftime("%Y%m%d"), daydiff = 0):
    try:
        connection = pymysql.connect(host='192.168.40.202', port=3306,
                                     user='trader', password='123456',
                                     db='MailSaver',
                                     charset='utf8')
    except Exception as e:
        logger.error("Sever connect fail: {0}".format(e))
        return (False, [])
    else:
        cursor = connection.cursor()
        if daydiff > 0:
            sql = "SELECT DATE_FORMAT(date, \"%Y%m%d\") date from Tradedate WHERE date >= STR_TO_DATE(\"{0}\", \"%Y%m%d\") ORDER BY date LIMIT {1}, 1;".format(date, daydiff - 1)
        elif daydiff < 0:
            sql = "SELECT DATE_FORMAT(date, \"%Y%m%d\") date from Tradedate WHERE date <= STR_TO_DATE(\"{0}\", \"%Y%m%d\") ORDER BY date DESC LIMIT {1}, 1;".format(date, daydiff * (-1))
        else:
            sql = "SELECT DATE_FORMAT(date, \"%Y%m%d\") date from Tradedate WHERE date >= STR_TO_DATE(\"{0}\", \"%Y%m%d\") ORDER BY date LIMIT 1;".format(date)
        try:
            cursor.execute(sql)
            data = cursor.fetchone()
            return data[0]
        except Exception as e:
            raise e

    return 0


# 从字符串中提取日期，格式的也是要是“YYYYMMDD”或者“YYYY年MM/M月DD/D日”
# 提取成功，返回字符串YYYYMMDD，否则返回None
def GetDateByString(source_str):
    result = None
    re_str = '(\D|$)*\d{8}(\D|$)*'
    match = re.search(re_str, source_str)
    if match:
        re_str = '\d+'
        result = ''.join(re.findall(re_str, match.group(0)))
    else:
        re_str = '(\d{4})年(\d{1,2})月(\d{1,2})日'
        match = re.search(re_str, source_str)
        if match:
            re_str = '\d+'
            datelist = re.findall(re_str, match.group(0))
            result = datelist[0].zfill(4) + datelist[1].zfill(2) + datelist[2].zfill(2)
    return result


# 把数字转换成日期，比如43406，实际是18991230后面加43406天
# 返回YYYYMMDD的字符串
def conver_number2date(numbers):
    delta = datetime.timedelta(days=numbers)
    today = datetime.datetime.strptime('18991230', '%Y%m%d') + delta
    return datetime.datetime.strftime(today, '%Y%m%d')


# 返回本机的MAC
def GetLocalMac(separater=':'):
    mac = uuid.UUID(int=uuid.getnode()).hex[-12:]
    return (separater.join([mac[e:e + 2] for e in range(0, 11, 2)])).upper()


# 返回本机的MAC，是个字典，本地回环地址会被过滤掉
# 返回值：{'以太网 2': '00-FF-42-69-29-68', 'WLAN': '74-E5-F9-DE-C1-5A'}
def get_local_mac_dict(is_up=True):
    up_if_list = []
    if is_up:
        status_dic = psutil.net_if_stats()
        key_list = status_dic.keys()
        for if_key in key_list:
            if_status = status_dic[if_key]
            if if_status.isup:
                up_if_list.append(if_key)

    ori_if_dic = psutil.net_if_addrs()
    key_list = ori_if_dic.keys()
    out_if_info_dic = {}
    for if_key in key_list:
        if if_key.lower().find('loopback') >= 0:
            continue
        if is_up:
            if if_key in up_if_list:
                if_info_list = ori_if_dic[if_key]
                out_if_info_dic[if_key] = if_info_list[0].address
        else:
            if_info_list = ori_if_dic[if_key]
            out_if_info_dic[if_key] = if_info_list[0].address
    return out_if_info_dic


# 解压zip文件
# filename:     解压的文件名，绝对路径
# path:         解压到的路径
# passwordlist: 密码列表，解压可能会用到密码
# 无返回值，文件解压在当前路径下
def unzipfile(filename, path = None, passwordlist = None):
    global logger
    dir_path = os.path.dirname(filename)
    unzip_result = False
    if filename.rfind(".zip") > 0:
        if logger:
            logger.info("Unzip file " + filename)
        try:
            zip_files = zipfile.ZipFile(filename)
            if path:
                zip_files.extractall(path, zip_files.namelist())
            else:
                zip_files.extractall(dir_path, zip_files.namelist())
            unzip_result = True
            zip_files.close()
        except RuntimeError as e:
            if passwordlist:
                for password in passwordlist:
                    if logger:
                        logger.debug("Try password {0} to unzip file {1}".format(password, filename))
                    try:
                        zip_files.extractall(dir_path, zip_files.namelist(), password.encode('utf-8'))
                        unzip_result = True
                        zip_files.close()
                    except Exception as e:
                        if logger:
                            logger.error("Unzip {0} failed. Bad pasword {1}!".format(filename, password))
        except Exception as e2:
            logger.error("Unzip {0} failed. Exception occured: {1}".format(filename, e2))
        if not unzip_result:
            if logger:
                logger.info("Unzip file {0} failed!".format(filename))


# 解压rar文件
# zfile:    解压的文件名，绝对路径
# path:     解压到的路径
# 无返回值，失败会raise Exception
def unrarfile(zfile, path):
    if not os.path.exists(path):
        os.mkdir(path)
    rar_command1 = "WinRAR.exe x -ibck -y %s %s" % (zfile, path)
    rar_command2 = r'"C:\Program Files\WinRAR\WinRAR.exe" x -ibck -y %s %s' % (zfile, path)
    if os.system(rar_command1) != 0:
        if os.system(rar_command2) != 0:
            raise Exception('Unrar {0} failed!'.format(zfile))

########################################################################################################################
## Mail
########################################################################################################################


MAX_MAIL_COUNT = 500


def decode_str(s):
    # decode_header()返回一个list，因为像Cc、Bcc这样的字段可能包含多个邮件地址，所以解析出来的会有多个元素。
    # 这里只取了第一个元素。如果是解析To字段，那么这种只分析第一个字段的做法是有问题的
    value, charset = decode_header(s)[0]
    if charset:
        value = value.decode(charset)
    return value


# 解析主题
def parse_subject(msg):
    value = msg.get('Subject', '')
    value = decode_str(value)
    return value


# 解析收件人，返回list
def parse_to(msg):
    return parseaddr(msg.get("To"))


# 解析发件人，返回list
def parse_from(msg):
    return parseaddr(msg.get("From"))


def SendingMailFromIT(Receive_list= [], Subject= '', Text= '', Attachpath= None):
    """ Send mail with attaching files from sender[IT]

    Sending mail server get from mysql[192.168.40.202/MailSaver] with user trader/123456

    Args:
        Receive_list： mail receivers list
        Subject: mail subject
        Text: mail text
        Attachpath: if there is an attach, here is filepath

    Returns:
        None

    Raises:
        Mysql error: Mysql execute error
        Mailserver error: mail server connect/sending error
    """
    global logger
    try:
        #数据库连接
        connection = pymysql.connect(host='192.168.40.202', port=3306,\
            user='trader',password='123456',db='MailSaver',charset='utf8')
    except Exception as e:
        if logger:
            logger.error("Sever connect fail: {0}".format(e))
    else:
        cursor = connection.cursor()
        sql = "SELECT ts.id, ts.Address, ts.ServerName, ts.Username, ts.`Password`, ts.Use_SSL, ts.port from ServerInfo ts \
            WHERE ServerName = \'IT\' AND type = \'SMTP\';"
        try:
            #对接发送邮件服务器
            if logger:
                logger.debug('select sql: ' + sql)
            cursor.execute(sql)
            serverdata = cursor.fetchall()
            if logger:
                logger.debug("Get mail server info:%s" % serverdata)
        except Exception as e:
            if logger:
                logger.error("Sql execute error: {0}".format(e))

        # 关闭数据库
        cursor.close()
        connection.close()

    if serverdata:
        host = serverdata[0][1]
        port = serverdata[0][6]
        is_ssl = serverdata[0][5]
        username = serverdata[0][3]
        password = serverdata[0][4]

        msg = MIMEMultipart()
        msg['From'] = username
        msg['To'] = ','.join(Receive_list)
        if Subject == '':
            if Attachpath:
                filename = os.path.basename(Attachpath)
                logger.debug(Attachpath)
                logger.debug(filename)
                Subject = filename
        msg['Subject'] = Header(Subject, 'utf-8').encode()
        # 邮件正文是MIMEText:
        msg.attach(MIMEText(Text, 'plain', 'utf-8'))

        #不知道类型的文件，都是application/octet-stream
        if Attachpath:
            [dirname, filename] = os.path.split(Attachpath)
            part = MIMEApplication(open(Attachpath, 'rb').read())
            part.add_header('Content-Disposition', 'attachment', filename=filename)
            # 添加到MIMEMultipart:
            msg.attach(part)
        if logger:
            logger.debug('Sending mail info: '.format(msg.as_string()))

        try:
            if is_ssl == '1':
                server = smtplib.SMTP_SSL(host, port)
            else:
                server = smtplib.SMTP(host, port)
            server.login(username, password)
            server.sendmail(username, Receive_list, msg.as_string())
            if logger:
                if Attachpath:
                    logger.info("Mail subject[{0}] attachment[{1}] sending to {2}".format(Subject, Attachpath, Receive_list))
                else:
                    logger.info("Mail subject[{0}] sending to {1}".format(Subject, Receive_list))
            server.quit()  # 关闭连接
        except Exception as e:
            if logger:
                logger.error("Send mail error: {0}".format(e))
    else:
        if logger:
            logger.info("Get mail server info error, no server!")

# 连接Settlement邮箱，返回服务器对象
# 返回(0, e)或者(1, Server)
def ConnectSettmentMailServer():
    global logger
    try:
        connection = pymysql.connect(host='192.168.40.202', port=3306,\
            user='trader',password='123456',db='MailSaver',charset='utf8')
    except Exception as e:
        if logger:
            logger.error("Sever connect fail: {0}".format(e))
        return (0, e)
    else:
        cursor = connection.cursor()
        sql_server = "SELECT ts.id, ts.Address, ts.ServerName, ts.Username, ts.`Password`, ts.Use_SSL, ts.port from ServerInfo ts WHERE ServerName = \'Settlement\';"
        cursor.execute(sql_server)
        serverdata = cursor.fetchall()
        cursor.close()
        connection.close()

        servername = serverdata[0][2]
        address = serverdata[0][1]
        port = serverdata[0][6]
        is_ssl = serverdata[0][5]
        username = serverdata[0][3]
        password = serverdata[0][4]
        if logger:
            logger.debug("Connect to {0}, address is {1}:{2}, username is {3}".format(servername, address, port, username))
        try:
            if is_ssl:
                if logger:
                    logger.debug("Connect to server using SSL.")
                server = poplib.POP3_SSL(address, port=port)
            else:
                server = poplib.POP3(address, port=port)
            server.user(username)
            server.pass_(password)
            # stat()返回邮件数量和占用空间:
            if logger:
                logger.info('Messages: %s. Size: %s' % server.stat())
        except Exception as e:
            if logger:
                logger.error("Server connect fail: %s" % e)
            return (0, e)
        else:
            return (1, server)


# 对接邮箱，从中按照SubjectMatchTable筛选主题，保存附件
# Server：               邮箱连接对象
# SubjectMatchTable：    主题匹配表，Dataframe，列：
#       Key：                关键字
#       Subject：            匹配的主题关键字
#       SubjectMatchMode：   主题匹配模式：1-从头；2-从中间；3-从末尾
#       Sender：             发件人，非必须
#       Path：               附件保存位置
#       AttachSaveMode：     附件保存路径文件夹名字模式：1-用日期作为文件夹名；2-用关键字作为文件夹名
# start：                邮件检索的起始位置，位置1代表最早的邮件，默认是0，表示从最后500封开始
# MaxMailCount：         最大邮件检索数，默认是MAX_MAIL_COUNT=500
# Unzip：                附件是否解压缩
# ZipPasswordList：      附件如果解压缩且需要密码，解压缩密码序列
# 返回有保存附件的Key和对应的日期，Dataframe
#       Key
#       Savedate
def ParseMailSaveAttach(Server, SubjectMatchTable, Start=0, MaxMailCount=MAX_MAIL_COUNT, Unzip=False,
                        ZipPasswordList=[]):
    if 'Sender' in SubjectMatchTable.columns:
        check_sender = True
    else:
        check_sender = False

    global logger
    result_list = []
    # 可能邮件正文太长，需要扩大，默认是2048
    poplib._MAXLINE = 20480
    # list()返回所有邮件的编号:
    resp, mails, octets = Server.list()
    # 默认取最后500封，如果有起始位置，就用起始位置
    if Start == 0:
        Start = len(mails) - MaxMailCount
    # 注意邮件的索引号从1开始
    for index in range(Start, len(mails) + 1):
        if logger:
            logger.debug("Parsing mail[{0}]...".format(index))
        # 解析这里有可能会出莫名其妙的错误，主要是从服务器可能断开了连接，要重新连上去继续读取
        try:
            # 这里如果不扩展poplib._MAXLINE，可能会返回 line too long 错误，就是返回的lines长度超过了界限
            resp, lines, octets = Server.retr(index)
        except Exception as e:
            if logger:
                logger.error("Parse mail[{0}] error: {1}".format(index, e))
            return (1, result_list, index)

        # lines存储了邮件的原始文本的每一行,
        # 可以获得整个邮件的原始文本:
        try:
            msg_content = b'\r\n'.join(lines).decode('utf-8')
        except UnicodeDecodeError:
            try:
                msg_content = b'\r\n'.join(lines).decode('GBK')
            except UnicodeDecodeError:
                try:
                    msg_content = b'\r\n'.join(lines).decode('GB2312')
                except Exception as e:
                    if logger:
                        logger.error('mail content decode error:{0}'.format(e))
                        logger.debug('decode mail fail mail id {0}, mail context: {1}'.format(index, lines))
                        logger.info('skip mail[{0}]'.format(index))
                    continue
        # 稍后解析出邮件:
        msg = Parser().parsestr(msg_content)
        subject = parse_subject(msg)
        from_list = parse_from(msg)
        to_list = parse_to(msg)
        logger.debug("Parsing mail[{0}] subject[{1}] from[{2}] to[{3}]...".format(index, subject, from_list, to_list))

        # 根据SubjectMatchTable过滤主题
        for index, row in SubjectMatchTable.iterrows():
            Key = row['Key']
            Subject = row['Subject']
            SubjectMatchMode = row['SubjectMatchMode']
            Path = row['Path']
            AttachSaveMode = row['AttachSaveMode']
            if check_sender:
                sender = row['Sender']
                if sender not in from_list:
                    continue

            if StringMatch(SubjectMatchMode, Subject, subject):
                # 解析日期字段，先从主题中解析，如果不行，需要从附件名中解析
                save_date = GetDateByString(subject)

                for part in msg.walk():
                    filename = part.get_filename()
                    if filename:
                        attachfilename = decode_str(filename)
                        logger.info("Attachment name: {0}".format(attachfilename))
                        data = part.get_payload(decode=True)
                        # 解析日期字段，主题中如果没有日期，需要从附件名中解析
                        if save_date is None:
                            save_date = GetDateByString(attachfilename)
                            if save_date is None:
                                save_date = datetime.datetime.now().strftime("%Y%m%d")

                        if AttachSaveMode == 1:
                            # 以日期为子文件夹名
                            save_path = os.path.join(Path, save_date)
                        else:
                            # 以关键字为子文件夹名
                            save_path = os.path.join(Path, Key)
                        if not os.path.exists(save_path):
                            os.mkdir(save_path)
                        logger.info("Save file[{0}] to {1}".format(attachfilename, save_path))
                        w_file = open(os.path.join(save_path, attachfilename), "wb")
                        w_file.write(data)
                        w_file.close()
                        result_list.append([Key, save_date])

                        [filename_prefix, filename_suffix] = attachfilename.split('.')
                        if attachfilename.lower().rfind(".zip") > 0 and Unzip:
                            unzipfile(os.path.join(save_path, attachfilename), passwordlist=ZipPasswordList)
                        elif attachfilename.lower().rfind('.rar') > 0 and Unzip:
                            unrarfile(os.path.join(save_path, attachfilename), os.path.join(save_path, filename_prefix))
            else:
                logger.debug("mail[{0}] subject[{1}] unmatch[{2}][{3}].".format(index, subject, SubjectMatchMode, Subject))

    return (0, result_list, 0)

########################################################################################################################
## Excel
########################################################################################################################

def WriteData2Excel(data, filepath):
    """ Write data to specified path with Excel(.xlsx) format

    Data shoud be more than one row. Filepath is absolute path for file saving. Excel is .xlsx format

    Args:
        data: list, should be rows
        filepath: file sava to an absolute path

    Returns:
        None

    Raises:
        ExcelError: excel data proceed error
        IOError: IO error
    """
    if len(data) > 0:
        global logger
        wb = openpyxl.Workbook()
        ws = wb.active

        '''		
        i = 0
        for i in range(0, len(data)):
            row = data[i]
            for j in range(0, len(row)):
                print(row[j])
                ws.cell(row=i + 1, column=j + 1).value = row[j]
        '''

        for row in data:
            ws.append(row)
        wb.save(filepath)
        if logger:
            logger.info("WriteData2Excel[{0}]".format(filepath))

def WriteMultiData2Excel(I_filepath, *I_data):
    """ Write data to specified path with Excel(.xlsx) format, more than one sheet

    Dataname is the namelist of sheets. Data shoud be more than one row. Filepath is absolute path for file saving. Excel is .xlsx format

    Args:
        I_filepath: file sava to an absolute path
        I_data: format: (namelist, datalist) namelist is sheet name list and datalist is data in each sheet. In sheet, each data should be rows

    Returns:
        None

    Raises:
        ExcelError: excel data proceed error
        IOError: IO error
    """

    if len(I_data) > 0:
        global logger
        wb = openpyxl.Workbook();
        ws = wb.active
        namelist = I_data[0]
        datalist = I_data[1]
        for i in range(0, len(namelist)):
            ws.title = namelist[i]
            for row in datalist[i]:
                ws.append(row)
            ws = wb.create_sheet()
        wb.save(I_filepath)
        if logger:
            logger.info("WriteMultiData2Excel[{0}]".format(I_filepath))

def ReadFromExcel(I_filepath):
    """ Read data from specified filepath with Excel(.xlsx) format
    First row is description. Filepath is an absolute path with filename, either '.xlsx' or '.xls'.

    Args:
        data: list, should be rows
        I_filepath: an absolute filepath with file extension '.xlsx' or '.xls'

    Returns:
        If there is a file error or no data, return empty list []
        datalist format [row1, row2, row3...]
        row format [dict1, dict2, dict3...]
        dict format: data_type : data_value
        data type: text, number, empty
    Raises:
        No exception. Error are logged.
    """
    global logger
    if not os.path.isfile(I_filepath):
        logger.error('Filepathp[{0}] is not a file!'.format(I_filepath))
        return []
    fullfilename = os.path.basename(I_filepath)
    (filename,extension) = os.path.splitext(fullfilename)
    if extension.upper() != '.XLSX' and extension.upper() != '.XLS':
        logger.error('File[{0}] is not a EXCEL file!'.format(fullfilename))
        return []

    if extension.upper() == '.XLSX':
        wb = openpyxl.load_workbook(I_filepath)
        sheet = wb.active
        data = []
        for i in range(sheet.min_row, sheet.max_row + 1):
            one_row = []
            for j in range(sheet.min_column, sheet.max_column + 1):
                #这里把None单独拉出来，是因为openpyxl里面 TYPE_NUMERIC 和 TYPE_NULL 的值一样，都是'n'
                if sheet.cell(i, j).value == None:
                    s_type = 'empty'
                    onedata = ''
                else:
                    if sheet.cell(i, j).data_type == sheet.cell(i, j).TYPE_STRING:
                        s_type = 'text'
                    elif sheet.cell(i, j).data_type == sheet.cell(i, j).TYPE_NUMERIC:
                        s_type = 'number'
                    else:
                        s_type = sheet.cell(i, j).data_type
                    onedata = sheet.cell(i, j).value
                one_row.append({s_type:onedata})
            data.append(one_row)
    else:
        wb = xlrd.open_workbook(I_filepath)
        sheet = wb.sheets()[0]
        data = []
        for i in range(0, sheet.nrows):
            one_row = []
            for j in range(0, sheet.ncols):
                if sheet.cell(i, j).ctype == xlrd.XL_CELL_EMPTY:
                    s_type = 'empty'
                elif sheet.cell(i, j).ctype == xlrd.XL_CELL_TEXT:
                    s_type = 'text'
                elif sheet.cell(i, j).ctype == xlrd.XL_CELL_NUMBER:
                    s_type = 'number'
                else:
                    s_type = sheet.cell(i, j).ctype
                one_row.append({s_type:sheet.cell(i, j).value})
            data.append(one_row)

    return data

