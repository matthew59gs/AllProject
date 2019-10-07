#!/usr/bin/env python
# -*- coding:utf-8 -*-
# 从数据库里面导出产品的持仓，每日一个excel

import pymysql
from openpyxl import *
import os

#建立一个以产品名字命名的文件夹，后面生产的Excel，都放到应对产品名字的文件夹中去
def MakeFolder(product_name):
	path = os.getcwd()
	if not os.path.exists(path + os.sep + product_name):
		os.mkdir(path + os.sep + product_name)
	return path + os.sep + product_name

#检索每一行的第一个日期字段，如果是同一个日期，要保存在一个Excel里，不同日期的数据，用不同的Excel保存
#fields里面只是列名，data里面不含列名
def SaveData2Excel(product_name, fields, data, path):
	savedate = data[0][0];
	wb = Workbook()
	ws = wb.active
	#数据的行号从2开始，第1行是列名
	ws.append([f[0] for f in fields])
	row_num = 2
	for row in data:
		rowdate = row[0]
		if savedate != rowdate:
			wb.save(path + os.sep + str(savedate) + "_" + product_name + "_持仓.xlsx")
			savedate = rowdate
			wb = Workbook()
			ws = wb.active
			#第一行是列名
			ws.append([f[0] for f in fields])
			row_num = 2
		else:
			for col in range(0, len(fields)):
				_ = ws.cell(column = col + 1, row = row_num, value = u'%s'%row[col])
			row_num += 1
	wb.save(path + os.sep + str(savedate) + "_" + product_name + "_持仓.xlsx")

#需要导出数据的产品和对应的日期范围
productlist = {"白塔1号", "峰云1号", "之江1号", "之江2号"}
begin_date = 20180201
end_date = 20180430

#数据库地址
connection = pymysql.connect(host='192.168.40.202', port=3306,\
	user='oplus',password='oplus',db='oplus_p',charset='utf8')
cursor = connection.cursor()
#与title的值相互对应
sql_base = "SELECT\
	tf.business_date AS '交易日',\
	tf.fund_name AS '产品名称',\
	CASE \
WHEN tf.market_no = '1' THEN\
	'上交所' \
WHEN tf.market_no = '2' THEN\
	'深交所' \
WHEN tf.market_no = '35' THEN\
	'沪港通' \
WHEN tf.market_no = '36' THEN\
	'深港通' \
ELSE\
	tf.market_no \
END AS '市场名称',\
 tf.report_code AS '证券代码',\
 tf.stock_name AS '证券名称',\
 tf.current_amount AS '当前数量',\
 tf.current_cost AS '当前成本',\
 tf.stock_asset_value AS '当前市值',\
 tf.today_profit AS '总盈亏',\
 tf.floating_profit AS '浮动盈亏' \
FROM\
	tfundstock tf \
WHERE\
	tf.current_amount > 0 \
AND tf.fund_name LIKE '%{0}%' \
AND tf.business_date >= {1} \
AND tf.business_date <= {2};"

#从数据库获取数据
for product_name in productlist:
	sql = sql_base.format(product_name, begin_date, end_date)
	cursor.execute(sql)
	result = cursor.fetchall()
	fields = cursor.description
	path = MakeFolder(product_name)
	SaveData2Excel(product_name, fields, result, path)
	print('{0} is over'.format(product_name))

#关闭数据库
cursor.close()
connection.close()